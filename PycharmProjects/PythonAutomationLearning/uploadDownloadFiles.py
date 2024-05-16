import time
from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

def update_excel_data(filepath,searchTerm,colName,new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict = {}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value ==colName:
            Dict["col"]=i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row =i,column=j).value ==searchTerm:
                Dict["row"]=i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)



file_path = "C:/Users/USER/Downloads/download.xlsx"
fruit_name ='Apple'
newvalue="900"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.maximize_window()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.CLASS_NAME,"button").click()
time.sleep(20)

# Edit the excel with the updated value



update_excel_data(file_path,fruit_name,"price",newvalue)

# Upload the modified Excel sheet.
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver,10)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
pricecolumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+pricecolumn+"-undefined']").text
print("The price of "+fruit_name+ "is"+actual_price)
assert actual_price == newvalue



